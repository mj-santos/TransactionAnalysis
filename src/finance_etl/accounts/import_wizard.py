"""Account import wizard: CSV/XLSX detection, column mapping, type inference, commit.

Supports three import scenarios:
  1. Single file, single sheet (row range selector)
  2. Single file, multiple sheets (XLSX sheet assignment)
  3. Two separate files

Reuses Spendly's existing csv_sniff + wizard_mapping infrastructure.
"""
from __future__ import annotations

import csv
import io
import re
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

# ── Canonical fields for account import ──────────────────────────────

LIABILITY_FIELDS = [
    "account_name", "current_balance", "statement_balance", "minimum_payment",
    "due_date", "credit_limit", "interest_rate", "payment_source",
    "account_type", "institution", "last_four", "notes",
]

ASSET_FIELDS = [
    "account_name", "current_balance", "account_type", "institution",
    "payment_source_tag", "notes",
]

REQUIRED_FIELDS = {"account_name", "current_balance"}

# Keyword maps for auto-suggesting column → canonical field
_FIELD_KEYWORDS: dict[str, list[str]] = {
    "account_name": ["account", "name", "vendor", "payee", "creditor", "card", "description"],
    "current_balance": ["balance", "current", "owed", "outstanding", "amount"],
    "statement_balance": ["statement", "statementbalance", "amountdue", "billed"],
    "minimum_payment": ["minimum", "minpayment", "mindue", "minimumdue", "min"],
    "due_date": ["due", "duedate", "paymentdue", "payby", "day"],
    "credit_limit": ["limit", "creditlimit", "creditline"],
    "interest_rate": ["apr", "rate", "interest", "interestrate"],
    "payment_source": ["source", "payfrom", "paymentaccount", "fundedby", "pay from"],
    "account_type": ["type", "accounttype", "category"],
    "institution": ["bank", "issuer", "institution", "provider", "company"],
    "last_four": ["last4", "lastfour", "ending", "accountnumber", "last four"],
    "notes": ["notes", "memo", "comment"],
    "payment_source_tag": ["tag", "code", "shortcode", "alias"],
}

# Account type inference patterns
_TYPE_PATTERNS: list[tuple[str, str, str]] = [
    # (regex_pattern, account_class, subtype)
    (r"mortgage|home\s*loan", "liability", "mortgage"),
    (r"auto|car\s*loan|truck|bronco|subaru|toyota|honda|ford|chevy|vehicle", "liability", "auto_loan"),
    (r"student|school|navient|fedloan|mohela|nelnet", "liability", "student_loan"),
    (r"energ|electric|water|gas|verizon|phone|utility|power|cable|internet|comcast|att|tmobile|spectrum", "liability", "utility"),
    (r"checking|ch$|\bchk\b", "asset", "checking"),
    (r"savings|sv$|\bsav\b", "asset", "savings"),
    (r"invest|brokerage|401k|ira|roth", "asset", "investment"),
    (r"venmo|paypal|zelle|cashapp|apple\s*cash", "asset", "digital_wallet"),
]

# Default: if class is liability and nothing matched → credit_card
# Default: if class is asset and nothing matched → checking


def _now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


# ── File detection ───────────────────────────────────────────────────

def detect_file(file_path: str, max_preview_rows: int = 10) -> dict:
    """
    Detect file type, read headers and preview rows.
    For XLSX, also detect sheet names.
    Returns file metadata for the wizard UI.
    """
    path = Path(file_path)
    suffix = path.suffix.lower()

    if suffix in (".xlsx", ".xls"):
        return _detect_xlsx(path, max_preview_rows)
    else:
        return _detect_csv(path, max_preview_rows)


def _detect_csv(path: Path, max_rows: int) -> dict:
    """Detect CSV file: encoding, delimiter, headers, preview rows."""
    from finance_etl.utils.csv_sniff import sniff_csv

    profile = sniff_csv(str(path))
    headers = profile["headers"]
    encoding = profile["encoding"]
    delimiter = profile["delimiter"]

    all_rows = []
    try:
        with open(path, encoding=encoding, errors="replace", newline="") as f:
            reader = csv.reader(f, delimiter=delimiter)
            next(reader, None)  # skip header
            for i, row in enumerate(reader):
                all_rows.append(row)
    except Exception:
        pass

    # Auto-detect section boundaries (blank rows, total rows)
    boundaries = _detect_boundaries(all_rows, headers)

    preview_rows = all_rows[:max_rows]

    return {
        "file_type": "csv",
        "file_path": str(path),
        "encoding": encoding,
        "delimiter": delimiter,
        "headers": headers,
        "row_count": len(all_rows),
        "preview_rows": preview_rows,
        "boundaries": boundaries,
        "sheets": None,
    }


def _detect_xlsx(path: Path, max_rows: int) -> dict:
    """Detect XLSX file: sheet names, headers per sheet, previews."""
    try:
        import openpyxl
    except ImportError:
        return {
            "file_type": "xlsx",
            "file_path": str(path),
            "error": "openpyxl not installed. Install with: pip install openpyxl",
            "sheets": None,
        }

    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    sheets = {}
    for name in wb.sheetnames:
        ws = wb[name]
        rows = []
        headers = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            cells = [str(c) if c is not None else "" for c in row]
            if i == 0:
                headers = cells
            else:
                rows.append(cells)
            if i >= max_rows:
                break

        # Guess which type this sheet is
        sheet_type = _guess_sheet_type(name, headers)

        sheets[name] = {
            "headers": headers,
            "row_count": ws.max_row - 1 if ws.max_row else 0,
            "preview_rows": rows[:max_rows],
            "suggested_type": sheet_type,
        }
    wb.close()

    return {
        "file_type": "xlsx",
        "file_path": str(path),
        "sheets": sheets,
        "sheet_names": list(wb.sheetnames),
    }


def _guess_sheet_type(name: str, headers: list[str]) -> str | None:
    """Guess if a sheet contains liability or asset data based on name/headers."""
    combined = (name + " " + " ".join(headers)).lower()
    if any(kw in combined for kw in ["credit", "card", "loan", "liability", "bill", "debt"]):
        return "liability"
    if any(kw in combined for kw in ["bank", "checking", "savings", "asset", "investment"]):
        return "asset"
    return None


def _detect_boundaries(rows: list[list[str]], headers: list[str]) -> list[dict]:
    """
    Detect section boundaries in a single-sheet file.
    Looks for blank rows, total rows, and section header rows.
    """
    boundaries = []
    total_keywords = {"total", "sum", "subtotal", "grand total", "net", "balance"}

    for i, row in enumerate(rows):
        row_num = i + 2  # 1-indexed, after header row
        text = " ".join(str(c).strip() for c in row).strip().lower()

        if not text or all(c.strip() == "" for c in row):
            boundaries.append({"row": row_num, "type": "blank"})
        elif any(kw in text for kw in total_keywords):
            boundaries.append({"row": row_num, "type": "total", "text": text[:80]})

    return boundaries


# ── Column mapping ───────────────────────────────────────────────────

def suggest_account_mappings(headers: list[str], section_type: str = "liability") -> dict[str, str | None]:
    """
    Suggest column → canonical field mappings using keyword matching.
    Returns {canonical_field: best_csv_header_or_None}.
    """
    fields = LIABILITY_FIELDS if section_type == "liability" else ASSET_FIELDS
    result: dict[str, str | None] = {f: None for f in fields}
    used: set[str] = set()

    # Normalize headers for matching
    norm_headers = {h: re.sub(r"[^a-z0-9]", "", h.lower()) for h in headers}

    for field in fields:
        keywords = _FIELD_KEYWORDS.get(field, [])
        for header in headers:
            if header in used:
                continue
            nh = norm_headers[header]
            if any(kw in nh for kw in keywords):
                result[field] = header
                used.add(header)
                break

    return result


# ── Type inference ───────────────────────────────────────────────────

def infer_account_type(name: str, default_class: str = "liability") -> tuple[str, str]:
    """
    Infer account_class and subtype from account name.
    Returns (account_class, subtype).
    """
    name_lower = name.lower()
    for pattern, acct_class, subtype in _TYPE_PATTERNS:
        if re.search(pattern, name_lower):
            return acct_class, subtype

    if default_class == "liability":
        return "liability", "credit_card"
    return "asset", "checking"


# ── Preview (pre-commit) ────────────────────────────────────────────

def preview_import(
    rows: list[list[str]],
    headers: list[str],
    mapping: dict[str, str],
    section_type: str = "liability",
    conn=None,
) -> dict:
    """
    Build a structured preview of what will be imported.
    Validates data, infers types, detects duplicates.
    """
    accounts = []
    warnings = []
    header_idx = {h: i for i, h in enumerate(headers)}

    def _get(row, canonical_field):
        csv_header = mapping.get(canonical_field)
        if not csv_header or csv_header not in header_idx:
            return None
        idx = header_idx[csv_header]
        return row[idx].strip() if idx < len(row) and row[idx] else None

    # Get existing account names for duplicate detection
    existing_names = set()
    if conn:
        try:
            existing = conn.execute("SELECT LOWER(name) FROM nw_accounts").fetchall()
            existing_names = {r[0] for r in existing}
        except Exception:
            pass

    for i, row in enumerate(rows):
        name = _get(row, "account_name")
        if not name:
            continue

        balance_str = _get(row, "current_balance")
        balance = _parse_number(balance_str) if balance_str else 0

        acct_class, subtype = infer_account_type(name, section_type)

        entry = {
            "row_index": i,
            "account_name": name,
            "current_balance": balance,
            "account_class": acct_class,
            "inferred_type": subtype,
            "statement_balance": _parse_number(_get(row, "statement_balance")),
            "minimum_payment": _parse_number(_get(row, "minimum_payment")),
            "due_date": _get(row, "due_date"),
            "credit_limit": _parse_number(_get(row, "credit_limit")),
            "interest_rate": _parse_number(_get(row, "interest_rate")),
            "payment_source": _get(row, "payment_source") or _get(row, "payment_source_tag"),
            "institution": _get(row, "institution"),
            "last_four": _get(row, "last_four"),
            "notes": _get(row, "notes"),
            "is_duplicate": name.lower() in existing_names,
        }
        accounts.append(entry)

    duplicates = [a for a in accounts if a["is_duplicate"]]
    if duplicates:
        warnings.append(f"{len(duplicates)} account(s) already exist: {', '.join(d['account_name'] for d in duplicates[:5])}")

    no_source = [a for a in accounts if not a.get("payment_source") and a["account_class"] == "liability"]
    if no_source:
        warnings.append(f"{len(no_source)} liability account(s) have no payment source assigned")

    return {
        "accounts": accounts,
        "total": len(accounts),
        "duplicates": len(duplicates),
        "warnings": warnings,
        "section_type": section_type,
    }


def _parse_number(val: str | None) -> float | None:
    """Parse a number from a string, handling currency symbols and commas."""
    if not val:
        return None
    # Remove currency symbols, commas, whitespace
    cleaned = re.sub(r"[$€£,\s]", "", val.strip())
    # Handle parentheses for negative numbers
    if cleaned.startswith("(") and cleaned.endswith(")"):
        cleaned = "-" + cleaned[1:-1]
    try:
        return round(float(cleaned), 2)
    except (ValueError, TypeError):
        return None


# ── Commit import ────────────────────────────────────────────────────

def commit_import(
    conn,
    accounts: list[dict],
    duplicate_action: str = "skip",
) -> dict:
    """
    Commit the import: create accounts and initial balance ledger entries.

    duplicate_action: 'skip' | 'update' | 'create'
      - skip: ignore accounts that already exist
      - update: update balance on existing accounts
      - create: create new accounts even if name matches
    """
    from .crud import create_account, get_account, list_accounts
    from .balance_ops import bulk_balance_update
    from .schemas import AccountCreate

    now = _now_iso()
    created = 0
    updated = 0
    skipped = 0

    # Build name→id lookup for existing accounts
    existing = list_accounts(conn)
    name_to_id = {a["name"].lower(): a["id"] for a in existing}

    for acct in accounts:
        name = acct["account_name"]
        name_lower = name.lower()
        balance = acct.get("current_balance", 0) or 0

        if name_lower in name_to_id and duplicate_action == "skip":
            skipped += 1
            continue

        if name_lower in name_to_id and duplicate_action == "update":
            acct_id = name_to_id[name_lower]
            bulk_balance_update(conn, [{
                "account_id": acct_id,
                "current_balance": balance,
                "statement_balance": acct.get("statement_balance"),
                "minimum_payment": acct.get("minimum_payment"),
                "data_source": "csv_import",
            }])
            # Update additional fields
            updates = {}
            if acct.get("credit_limit") is not None:
                updates["credit_limit"] = acct["credit_limit"]
            if acct.get("interest_rate") is not None:
                updates["interest_rate"] = acct["interest_rate"]
            if acct.get("institution"):
                updates["institution"] = acct["institution"]
            if updates:
                set_parts = ", ".join(f"{k} = ?" for k in updates)
                params = list(updates.values()) + [now, acct_id]
                conn.execute(
                    f"UPDATE nw_accounts SET {set_parts}, updated_at = ? WHERE id = ?",
                    params,
                )
            updated += 1
            continue

        # Create new account
        acct_class = acct.get("account_class", "liability")
        subtype = acct.get("inferred_type", "credit_card" if acct_class == "liability" else "checking")

        due_day = None
        if acct.get("due_date"):
            try:
                due_day = int(re.sub(r"[^0-9]", "", str(acct["due_date"])))
                if due_day < 1 or due_day > 31:
                    due_day = None
            except (ValueError, TypeError):
                due_day = None

        create_data = AccountCreate(
            name=name,
            account_class=acct_class,
            liability_type=subtype if acct_class == "liability" else None,
            asset_type=subtype if acct_class == "asset" else None,
            institution=acct.get("institution"),
            last_four=acct.get("last_four"),
            balance=abs(balance),
            credit_limit=acct.get("credit_limit"),
            interest_rate=acct.get("interest_rate"),
            due_day=due_day,
            statement_balance=acct.get("statement_balance"),
            payment_source_tag=acct.get("payment_source"),
            notes=acct.get("notes"),
        )

        try:
            result = create_account(conn, create_data)
            name_to_id[name_lower] = result["id"]
            created += 1
        except Exception:
            skipped += 1

    return {
        "created": created,
        "updated": updated,
        "skipped": skipped,
        "total": created + updated + skipped,
    }


# ── Read rows from file sections ─────────────────────────────────────

def read_csv_rows(
    file_path: str,
    start_row: int | None = None,
    end_row: int | None = None,
) -> tuple[list[str], list[list[str]]]:
    """
    Read CSV rows from a file, optionally slicing by row range.
    Row numbers are 1-indexed (row 1 = first data row after header).
    Returns (headers, rows).
    """
    from finance_etl.utils.csv_sniff import sniff_csv

    profile = sniff_csv(file_path)
    headers = profile["headers"]
    encoding = profile["encoding"]
    delimiter = profile["delimiter"]

    all_rows = []
    with open(file_path, encoding=encoding, errors="replace", newline="") as f:
        reader = csv.reader(f, delimiter=delimiter)
        next(reader, None)  # skip header
        for row in reader:
            all_rows.append(row)

    if start_row is not None and end_row is not None:
        # Convert 1-indexed to 0-indexed
        sliced = all_rows[start_row - 1:end_row]
    else:
        sliced = all_rows

    return headers, sliced


def read_xlsx_sheet(file_path: str, sheet_name: str) -> tuple[list[str], list[list[str]]]:
    """Read all rows from a specific XLSX sheet. Returns (headers, rows)."""
    try:
        import openpyxl
    except ImportError:
        return [], []

    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    ws = wb[sheet_name]
    headers = []
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        cells = [str(c) if c is not None else "" for c in row]
        if i == 0:
            headers = cells
        else:
            if any(c.strip() for c in cells):  # skip blank rows
                rows.append(cells)
    wb.close()
    return headers, rows
